import openpyxl
from openpyxl import Workbook
import os

def estatistica():
    # Abre a planilha
    wb = openpyxl.load_workbook("dados.xlsx")
    sheet = wb["lucro"]
    
    
    
    # Define as células iniciais e finais dos dados
    primeira_linha = 2
    ultima_linha = sheet.max_row

    # Inicializa as variáveis para os totais
    custo_total_total = 0
    lucro_total = 0

    # Calcula o total de custo total e lucro
    for linha in range(primeira_linha, ultima_linha + 1):
        custo_total_total += sheet.cell(row=linha, column=7).value
        lucro_total += sheet.cell(row=linha, column=8).value

    # Calcula a margem de lucro total
    preco_venda_total = sum(sheet.cell(row=linha, column=3).value for linha in range(primeira_linha, ultima_linha + 1))
    
    margem_lucro_total = 0  # Declaração da variável e atribuição de valor padrão
    if preco_venda_total ==0:
        margem_de_lucro_total = 0
    else:
        margem_lucro_total = lucro_total / preco_venda_total * 100


    custo_total = float("{:.2f}".format(custo_total_total))
    lucro_liquido_total = float("{:.2f}".format(lucro_total))
    margem_de_lucro_total = float("{:.2f}".format(margem_lucro_total))

    lista_dados = [custo_total, lucro_liquido_total, margem_de_lucro_total]

    return lista_dados


# ----------------------------------------

def obter_dados_excel(nome_arquivo):
    wb = openpyxl.load_workbook(nome_arquivo)
    sheet = wb.active
    dados = []

    for row in sheet.iter_rows(min_row=2, values_only=True):
        dados.append(row)

    return dados


# ----------------------------------------

# salvar Produto
def salvar_produto(nome_produto,preco_compra,preco_venda,quantidade,custos_adicionais,custo_frete):

    # Perguntando ao usuário pelos dados
    nome_produto = nome_produto
    preco_compra = float(preco_compra)
    preco_venda = float(preco_venda)
    quantidade = int(quantidade)
    custos_adicionais = float(custos_adicionais)
    custo_frete = float(custo_frete)

    # Calculando o lucro
    custo_total = (preco_compra + custos_adicionais + custo_frete) * quantidade
    lucro = (preco_venda - preco_compra - custos_adicionais - custo_frete) * quantidade
    margem_lucro = lucro / (preco_venda * quantidade) * 100

    # Salvando os resultados em uma folha do Excel
    try:
        wb = openpyxl.load_workbook("dados.xlsx")
    except FileNotFoundError:
        wb = openpyxl.Workbook()
        wb.active.title = "lucro"
        wb.active.append(["Produto", "Preço de compra", "Preço de venda", "Quantidade", "Custos adicionais", "Custo médio do frete", "Custo total", "Lucro líquido", "Margem de lucro"])

    sheet = wb["lucro"]
    last_row = sheet.max_row + 1

    sheet.cell(row=last_row, column=1).value = nome_produto
    sheet.cell(row=last_row, column=2).value = preco_compra
    sheet.cell(row=last_row, column=3).value = preco_venda
    sheet.cell(row=last_row, column=4).value = quantidade
    sheet.cell(row=last_row, column=5).value = custos_adicionais
    sheet.cell(row=last_row, column=6).value = custo_frete
    sheet.cell(row=last_row, column=7).value = custo_total
    sheet.cell(row=last_row, column=8).value = lucro
    sheet.cell(row=last_row, column=9).value = margem_lucro

    wb.save("dados.xlsx")


# Deletar produto
def deletar_linha_por_nome(nome_produto, nome_planilha):
    # Carrega a planilha
    wb = openpyxl.load_workbook(nome_planilha)
    sheet = wb.active
    contador = 2

    # Percorre a coluna de nomes procurando pelo nome do produto
    for row in sheet.iter_rows(min_row=2, min_col=1, max_col=1, values_only=True):
        if str(row[0]) == nome_produto:
            # Obtém o número da linha e deleta a linha inteira
            linha = contador
            sheet.delete_rows(linha)
            break

        # incrementando o contador
        contador +=1

    # Salva as mudanças na planilha
    wb.save(nome_planilha)
    
    
def criar_planilha():
    # Cria um novo arquivo Workbook
    wb = Workbook()

    # Seleciona a planilha padrão
    planilha = wb.active

    # Renomeia a planilha para "Lucro"
    planilha.title = "lucro"

    # Define os nomes das colunas
    colunas = ["Produto", "Preço de compra", "Preço de venda", "Quantidade", "Custos adicionais", 
               "Custo médio do frete", "Custo total", "Lucro líquido", "Margem de lucro"]

    # Insere os nomes das colunas na primeira linha
    for coluna, valor in enumerate(colunas, start=1):
        planilha.cell(row=1, column=coluna, value=valor)

    # Salva o arquivo
    wb.save("dados.xlsx")
    
    

def verificar_arquivo():
    if os.path.exists("dados.xlsx"):
        return True
    else:
        return False

def criar_ou_abrir_arquivo():
    if verificar_arquivo():
        # Arquivo existe, abra-o
        wb = openpyxl.load_workbook("dados.xlsx")
        return wb
    else:
        # Arquivo não existe, crie um novo
        wb = Workbook()
        wb.active.title = "lucro"
        wb.active.append(["Produto", "Preço de compra", "Preço de venda", "Quantidade", "Custos adicionais", "Custo médio do frete", "Custo total", "Lucro líquido", "Margem de lucro"])
        wb.save("dados.xlsx")
        return wb

# Verifica se o arquivo existe ou cria um novo
wb = criar_ou_abrir_arquivo()

# O restante do seu código continua aqui, usando a variável wb para acessar o arquivo.
